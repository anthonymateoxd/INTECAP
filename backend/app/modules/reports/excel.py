from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from app.modules.reports.schemas import (
    CommissionReportFilters,
    CommissionReportRow,
    FuelReportFilters,
    FuelReportRow,
    MaintenanceReportFilters,
    MaintenanceReportRow,
    TripReportFilters,
    TripReportRow,
    VehicleReportFilters,
    VehicleReportRow,
)
from decimal import Decimal



VEHICLE_REPORT_HEADERS = [
    "Código de inventario",
    "Placa",
    "Marca",
    "Modelo",
    "Año",
    "Tipo de vehículo",
    "Tipo de combustible",
    "Kilometraje actual (km)",
    "Capacidad del tanque (gal)",
    "Estado",
    "Fecha de registro",
]


def build_vehicles_excel(
    *,
    rows: list[VehicleReportRow],
    filters: VehicleReportFilters,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Vehículos"

    worksheet["A1"] = "Reporte de vehículos"
    worksheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    start_text = (
        filters.start_date.strftime("%d/%m/%Y")
        if filters.start_date
        else "Sin límite"
    )

    end_text = (
        filters.end_date.strftime("%d/%m/%Y")
        if filters.end_date
        else "Sin límite"
    )

    status_text = {
        "ALL": "Todos",
        "ACTIVE": "Activos",
        "INACTIVE": "Inactivos",
    }[filters.status]

    worksheet["A2"] = (
        f"Período: {start_text} - {end_text}"
    )
    worksheet["A3"] = (
        f"Estado: {status_text}"
    )

    header_row = 5

    for column_index, header in enumerate(
        VEHICLE_REPORT_HEADERS,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_index, report_row in enumerate(
        rows,
        start=header_row + 1,
    ):
        values = [
            report_row.inventory_code,
            report_row.license_plate,
            report_row.brand,
            report_row.model,
            report_row.year,
            report_row.vehicle_type,
            report_row.fuel_type,
            float(report_row.current_odometer_km),
            (
                float(report_row.tank_capacity_gal)
                if report_row.tank_capacity_gal is not None
                else None
            ),
            report_row.status,
            report_row.created_at.replace(
                tzinfo=None
            ),
        ]

        for column_index, value in enumerate(
            values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    # Formatos numéricos.
    for row_index in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        worksheet.cell(
            row=row_index,
            column=8,
        ).number_format = '#,##0.0'

        worksheet.cell(
            row=row_index,
            column=9,
        ).number_format = '#,##0.000'

        worksheet.cell(
            row=row_index,
            column=11,
        ).number_format = "dd/mm/yyyy hh:mm"

    # Filtro sobre las columnas del reporte.
    worksheet.auto_filter.ref = (
        f"A{header_row}:"
        f"K{max(header_row, worksheet.max_row)}"
    )

    worksheet.freeze_panes = (
        f"A{header_row + 1}"
    )

    # Anchos legibles.
    widths = {
        1: 22,
        2: 15,
        3: 16,
        4: 18,
        5: 10,
        6: 20,
        7: 20,
        8: 22,
        9: 25,
        10: 14,
        11: 22,
    }

    for column_index, width in widths.items():
        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    output = BytesIO()
    workbook.save(output)

    return output.getvalue()


COMMISSION_REPORT_HEADERS = [
    "Solicitante",
    "Vehículo / placa",
    "Estado",
    "Inicio programado",
    "Fin programado",
    "Revisado por",
    "Fecha de revisión",
    "Fecha de creación",
]


def build_commissions_excel(
    *,
    rows: list[CommissionReportRow],
    filters: CommissionReportFilters,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Comisiones"

    worksheet["A1"] = "Reporte de comisiones"
    worksheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    start_text = (
        filters.start_date.strftime("%d/%m/%Y")
        if filters.start_date
        else "Sin límite"
    )

    end_text = (
        filters.end_date.strftime("%d/%m/%Y")
        if filters.end_date
        else "Sin límite"
    )

    worksheet["A2"] = (
        f"Período: {start_text} - {end_text}"
    )

    worksheet["A3"] = (
        f"Total de comisiones: {len(rows)}"
    )

    header_row = 5

    for column_index, header in enumerate(
        COMMISSION_REPORT_HEADERS,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_index, report_row in enumerate(
        rows,
        start=header_row + 1,
    ):
        values = [
            report_row.requester_name,
            report_row.vehicle,
            report_row.status,
            report_row.scheduled_start_at.replace(
                tzinfo=None
            ),
            report_row.scheduled_end_at.replace(
                tzinfo=None
            ),
            report_row.reviewer_name or "-",
            (
                report_row.reviewed_at.replace(
                    tzinfo=None
                )
                if report_row.reviewed_at is not None
                else None
            ),
            report_row.created_at.replace(
                tzinfo=None
            ),
        ]

        for column_index, value in enumerate(
            values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    for row_index in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        for column_index in (4, 5, 7, 8):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "dd/mm/yyyy hh:mm"

    worksheet.auto_filter.ref = (
        f"A{header_row}:"
        f"H{max(header_row, worksheet.max_row)}"
    )

    worksheet.freeze_panes = (
        f"A{header_row + 1}"
    )

    widths = {
        1: 25,
        2: 28,
        3: 16,
        4: 22,
        5: 22,
        6: 25,
        7: 22,
        8: 22,
    }

    for column_index, width in widths.items():
        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    output = BytesIO()
    workbook.save(output)

    return output.getvalue()



TRIP_REPORT_HEADERS = [
    "Fecha",
    "Vehículo / placa",
    "Piloto",
    "No. recorrido",
    "Origen",
    "Destino",
    "Odómetro inicial (km)",
    "Odómetro final (km)",
    "Kilómetros recorridos",
    "Saldo de tanque (%)",
    "Tipo de carretera",
]


def build_trips_excel(
    *,
    rows: list[TripReportRow],
    filters: TripReportFilters,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Recorridos"

    worksheet["A1"] = "Reporte de recorridos y kilómetros"
    worksheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    start_text = (
        filters.start_date.strftime("%d/%m/%Y")
        if filters.start_date
        else "Sin límite"
    )

    end_text = (
        filters.end_date.strftime("%d/%m/%Y")
        if filters.end_date
        else "Sin límite"
    )

    worksheet["A2"] = (
        f"Período: {start_text} - {end_text}"
    )

    total_km = sum(
        (row.distance_km for row in rows),
        Decimal("0"),
    )

    worksheet["A3"] = (
        f"Total de recorridos: {len(rows)}"
    )

    worksheet["A4"] = (
        f"Kilómetros recorridos: {total_km} km"
    )

    header_row = 6

    for column_index, header in enumerate(
        TRIP_REPORT_HEADERS,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_index, report_row in enumerate(
        rows,
        start=header_row + 1,
    ):
        values = [
            report_row.trip_date,
            report_row.vehicle,
            report_row.driver_name,
            report_row.sequence_number,
            report_row.origin,
            report_row.destination,
            float(report_row.odometer_start_km),
            float(report_row.odometer_end_km),
            float(report_row.distance_km),
            (
                float(report_row.tank_balance_percent)
                if report_row.tank_balance_percent is not None
                else None
            ),
            report_row.road_type,
        ]

        for column_index, value in enumerate(
            values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    for row_index in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        worksheet.cell(
            row=row_index,
            column=1,
        ).number_format = "dd/mm/yyyy"

        for column_index in (7, 8, 9):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = '0.0'

        worksheet.cell(
            row=row_index,
            column=10,
        ).number_format = '0.00'

    worksheet.auto_filter.ref = (
        f"A{header_row}:"
        f"K{max(header_row, worksheet.max_row)}"
    )

    worksheet.freeze_panes = (
        f"A{header_row + 1}"
    )

    widths = {
        1: 14,
        2: 28,
        3: 25,
        4: 15,
        5: 24,
        6: 24,
        7: 22,
        8: 22,
        9: 22,
        10: 21,
        11: 20,
    }

    for column_index, width in widths.items():
        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    output = BytesIO()
    workbook.save(output)

    return output.getvalue()


FUEL_REPORT_HEADERS = [
    "Fecha y hora de carga",
    "Solicitante",
    "Vehículo / placa",
    "Gasolinera",
    "Galones",
    "Monto (Q)",
]


def build_fuel_excel(
    *,
    rows: list[FuelReportRow],
    filters: FuelReportFilters,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Combustible"

    worksheet["A1"] = "Reporte de combustible"
    worksheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    start_text = (
        filters.start_date.strftime("%d/%m/%Y")
        if filters.start_date
        else "Sin límite"
    )

    end_text = (
        filters.end_date.strftime("%d/%m/%Y")
        if filters.end_date
        else "Sin límite"
    )

    worksheet["A2"] = (
        f"Período: {start_text} - {end_text}"
    )

    worksheet["A3"] = (
        f"Total de cargas: {len(rows)}"
    )

    header_row = 5

    for column_index, header in enumerate(
        FUEL_REPORT_HEADERS,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_index, report_row in enumerate(
        rows,
        start=header_row + 1,
    ):
        values = [
            report_row.loaded_at.replace(
                tzinfo=None
            ),
            report_row.requester_name,
            report_row.vehicle,
            report_row.gas_station,
            float(report_row.gallons),
            float(report_row.amount_q),
        ]

        for column_index, value in enumerate(
            values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    for row_index in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        worksheet.cell(
            row=row_index,
            column=1,
        ).number_format = "dd/mm/yyyy hh:mm"

        worksheet.cell(
            row=row_index,
            column=5,
        ).number_format = "0.000"

        worksheet.cell(
            row=row_index,
            column=6,
        ).number_format = '"Q" #,##0.00'

    worksheet.auto_filter.ref = (
        f"A{header_row}:"
        f"F{max(header_row, worksheet.max_row)}"
    )

    worksheet.freeze_panes = (
        f"A{header_row + 1}"
    )

    widths = {
        1: 23,
        2: 28,
        3: 30,
        4: 28,
        5: 15,
        6: 16,
    }

    for column_index, width in widths.items():
        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    output = BytesIO()
    workbook.save(output)

    return output.getvalue()




MAINTENANCE_REPORT_HEADERS = [
    "Fecha de mantenimiento",
    "Vehículo / placa",
    "Tipo",
    "Odómetro (km)",
    "Taller",
    "Costo (Q)",
    "No. de factura",
    "Descripción",
    "Servicios realizados",
    "Repuestos utilizados",
    "Próximo mantenimiento",
    "Próximo mantenimiento (km)",
]


def build_maintenance_excel(
    *,
    rows: list[MaintenanceReportRow],
    filters: MaintenanceReportFilters,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Mantenimiento"

    worksheet["A1"] = "Reporte de mantenimiento"
    worksheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    start_text = (
        filters.start_date.strftime("%d/%m/%Y")
        if filters.start_date
        else "Sin límite"
    )

    end_text = (
        filters.end_date.strftime("%d/%m/%Y")
        if filters.end_date
        else "Sin límite"
    )

    worksheet["A2"] = (
        f"Período: {start_text} - {end_text}"
    )

    worksheet["A3"] = (
        f"Total de mantenimientos: {len(rows)}"
    )

    header_row = 5

    for column_index, header in enumerate(
        MAINTENANCE_REPORT_HEADERS,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_index, report_row in enumerate(
        rows,
        start=header_row + 1,
    ):
        values = [
            report_row.maintenance_date,
            report_row.vehicle,
            report_row.maintenance_type,
            (
                float(report_row.odometer_km)
                if report_row.odometer_km is not None
                else None
            ),
            report_row.workshop or "-",
            (
                float(report_row.cost_q)
                if report_row.cost_q is not None
                else None
            ),
            report_row.invoice_number or "-",
            report_row.description or "-",
            report_row.services,
            report_row.parts,
            report_row.next_maintenance_date,
            (
                float(
                    report_row.next_maintenance_odometer_km
                )
                if (
                    report_row.next_maintenance_odometer_km
                    is not None
                )
                else None
            ),
        ]

        for column_index, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for row_index in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        worksheet.cell(
            row=row_index,
            column=1,
        ).number_format = "dd/mm/yyyy"

        worksheet.cell(
            row=row_index,
            column=4,
        ).number_format = "0.0"

        worksheet.cell(
            row=row_index,
            column=6,
        ).number_format = '"Q" #,##0.00'

        worksheet.cell(
            row=row_index,
            column=11,
        ).number_format = "dd/mm/yyyy"

        worksheet.cell(
            row=row_index,
            column=12,
        ).number_format = "0.0"

    worksheet.auto_filter.ref = (
        f"A{header_row}:"
        f"L{max(header_row, worksheet.max_row)}"
    )

    worksheet.freeze_panes = (
        f"A{header_row + 1}"
    )

    widths = {
        1: 22,
        2: 30,
        3: 18,
        4: 18,
        5: 28,
        6: 16,
        7: 20,
        8: 35,
        9: 35,
        10: 35,
        11: 22,
        12: 27,
    }

    for column_index, width in widths.items():
        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    output = BytesIO()
    workbook.save(output)

    return output.getvalue()
