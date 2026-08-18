from getpass import getpass
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app


ADMIN_DEFAULT = "grabriel1@miumg.edu.gt"
COLLAB_DEFAULT = "prueba@intecap.edu.gt"

admin_email = input(
    f"Email ADMIN [{ADMIN_DEFAULT}]: "
).strip() or ADMIN_DEFAULT
admin_password = getpass("Password ADMIN: ")

collab_email = input(
    f"Email COLLABORATOR [{COLLAB_DEFAULT}]: "
).strip() or COLLAB_DEFAULT
collab_password = getpass("Password COLLABORATOR: ")


def login(client, email, password, label):
    response = client.post(
        "/api/v1/auth/login",
        json={
            "institutional_email": email,
            "password": password,
        },
    )

    print(f"LOGIN {label}:", response.status_code)

    if response.status_code != 200:
        print(response.json())
        raise RuntimeError(
            f"No fue posible iniciar sesión como {label}."
        )

    return response.json()["access_token"]


with TestClient(app) as client:
    admin_token = login(
        client,
        admin_email,
        admin_password,
        "ADMIN",
    )

    collab_token = login(
        client,
        collab_email,
        collab_password,
        "COLLABORATOR",
    )

    admin_headers = {
        "Authorization": f"Bearer {admin_token}",
    }

    collab_headers = {
        "Authorization": f"Bearer {collab_token}",
    }

    params = {
        "start_date": "2026-01-01",
        "end_date": "2026-12-31",
        "status": "ACTIVE",
    }

    # 1. Excel como ADMIN.
    response = client.get(
        "/api/v1/reports/vehicles/excel",
        headers=admin_headers,
        params=params,
    )

    print(
        "EXCEL ADMIN:",
        response.status_code,
    )

    print(
        "EXCEL CONTENT TYPE:",
        response.headers.get("content-type"),
    )

    print(
        "EXCEL BYTES:",
        len(response.content),
    )

    if response.status_code == 200:
        workbook = load_workbook(
            BytesIO(response.content),
            read_only=True,
        )

        worksheet = workbook.active

        print(
            "EXCEL SHEET:",
            worksheet.title.encode(
                "unicode_escape"
            ).decode("ascii"),
        )

        print(
            "EXCEL PLACA:",
            worksheet["B6"].value,
        )

        print(
            "EXCEL ESTADO:",
            worksheet["J6"].value,
        )

    # 2. PDF como ADMIN.
    response = client.get(
        "/api/v1/reports/vehicles/pdf",
        headers=admin_headers,
        params=params,
    )

    print(
        "PDF ADMIN:",
        response.status_code,
    )

    print(
        "PDF CONTENT TYPE:",
        response.headers.get("content-type"),
    )

    print(
        "PDF SIGNATURE:",
        response.content[:5],
    )

    print(
        "PDF BYTES:",
        len(response.content),
    )

    # 3. COLLABORATOR no puede generar Excel.
    response = client.get(
        "/api/v1/reports/vehicles/excel",
        headers=collab_headers,
        params=params,
    )

    print(
        "EXCEL COLLABORATOR:",
        response.status_code,
    )

    # 4. COLLABORATOR no puede generar PDF.
    response = client.get(
        "/api/v1/reports/vehicles/pdf",
        headers=collab_headers,
        params=params,
    )

    print(
        "PDF COLLABORATOR:",
        response.status_code,
    )

    # 5. Rango de fechas inválido.
    response = client.get(
        "/api/v1/reports/vehicles/pdf",
        headers=admin_headers,
        params={
            "start_date": "2026-12-31",
            "end_date": "2026-01-01",
            "status": "ALL",
        },
    )

    print(
        "INVALID DATE RANGE:",
        response.status_code,
    )

    # 6. Estado inválido.
    response = client.get(
        "/api/v1/reports/vehicles/pdf",
        headers=admin_headers,
        params={
            "status": "OTRO",
        },
    )

    print(
        "INVALID STATUS:",
        response.status_code,
    )

    # 7. Filtro INACTIVE.
    response = client.get(
        "/api/v1/reports/vehicles/excel",
        headers=admin_headers,
        params={
            "status": "INACTIVE",
        },
    )

    print(
        "EXCEL INACTIVE:",
        response.status_code,
    )

    if response.status_code == 200:
        workbook = load_workbook(
            BytesIO(response.content),
            read_only=True,
        )

        worksheet = workbook.active

        print(
            "INACTIVE FIRST DATA ROW:",
            worksheet["B6"].value,
        )

print("VEHICLE REPORT HTTP TEST OK")
