from getpass import getpass
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app


ADMIN_DEFAULT = "grabriel1@miumg.edu.gt"
COLLAB_DEFAULT = "prueba@intecap.edu.gt"

admin_email = input(
    f"Email ADMIN [{ADMIN_DEFAULT}]: "
).strip() or ADMIN_DEFAULT
admin_password = getpass("Password ADMIN: ")

collab_email = input(
    f"Email COLLABORATOR [{COLLAB_DEFAULT}]: "
).strip() or COLLAB_DEFAULT
collab_password = getpass("Password COLLABORATOR: ")


def login(client, email, password, label):
    response = client.post(
        "/api/v1/auth/login",
        json={
            "institutional_email": email,
            "password": password,
        },
    )

    print(f"LOGIN {label}:", response.status_code)

    if response.status_code != 200:
        raise RuntimeError(f"Login {label} falló.")

    return response.json()["access_token"]


with TestClient(app) as client:
    admin_token = login(
        client,
        admin_email,
        admin_password,
        "ADMIN",
    )

    collab_token = login(
        client,
        collab_email,
        collab_password,
        "COLLABORATOR",
    )

    admin_headers = {
        "Authorization": f"Bearer {admin_token}",
    }

    collab_headers = {
        "Authorization": f"Bearer {collab_token}",
    }

    params = {
        "start_date": "2026-01-01",
        "end_date": "2026-12-31",
    }

    # Excel ADMIN
    response = client.get(
        "/api/v1/reports/commissions/excel",
        headers=admin_headers,
        params=params,
    )

    print("EXCEL ADMIN:", response.status_code)

    if response.status_code == 200:
        workbook = load_workbook(
            BytesIO(response.content),
            read_only=True,
        )

        worksheet = workbook.active

        admin_total = max(
            worksheet.max_row - 5,
            0,
        )

        print("EXCEL ADMIN TOTAL:", admin_total)
        print(
            "EXCEL ADMIN FIRST USER:",
            worksheet["A6"].value,
        )
        print(
            "EXCEL ADMIN FIRST STATUS:",
            worksheet["C6"].value,
        )

    # Excel COLLABORATOR
    response = client.get(
        "/api/v1/reports/commissions/excel",
        headers=collab_headers,
        params=params,
    )

    print(
        "EXCEL COLLABORATOR:",
        response.status_code,
    )

    if response.status_code == 200:
        workbook = load_workbook(
            BytesIO(response.content),
            read_only=True,
        )

        worksheet = workbook.active

        collaborator_total = max(
            worksheet.max_row - 5,
            0,
        )

        print(
            "EXCEL COLLABORATOR TOTAL:",
            collaborator_total,
        )

        names = [
            worksheet.cell(
                row=row_number,
                column=1,
            ).value
            for row_number in range(
                6,
                worksheet.max_row + 1,
            )
        ]

        print(
            "COLLABORATOR USERS:",
            names,
        )

    # PDF ADMIN
    response = client.get(
        "/api/v1/reports/commissions/pdf",
        headers=admin_headers,
        params=params,
    )

    print("PDF ADMIN:", response.status_code)
    print(
        "PDF ADMIN TYPE:",
        response.headers.get("content-type"),
    )
    print(
        "PDF ADMIN SIGNATURE:",
        response.content[:5],
    )

    # PDF COLLABORATOR
    response = client.get(
        "/api/v1/reports/commissions/pdf",
        headers=collab_headers,
        params=params,
    )

    print(
        "PDF COLLABORATOR:",
        response.status_code,
    )
    print(
        "PDF COLLABORATOR SIGNATURE:",
        response.content[:5],
    )

    # Rango inválido
    response = client.get(
        "/api/v1/reports/commissions/excel",
        headers=admin_headers,
        params={
            "start_date": "2026-12-31",
            "end_date": "2026-01-01",
        },
    )

    print(
        "INVALID DATE RANGE:",
        response.status_code,
    )

print("COMMISSION REPORT HTTP TEST OK")
